"""export_dashboard_data.py tests for mart CSV and Excel exports."""

from __future__ import annotations

import json
from pathlib import Path

import duckdb
import pandas as pd
import pytest
from openpyxl import load_workbook

import create_duckdb_database as db_setup
import export_dashboard_data as dashboard_exports
import generate_recommendations as recommendations
import load_to_duckdb as loader
from clean_avazu_ads import clean_avazu_ads
from clean_hillstrom_email import clean_hillstrom_email
from helpers import (
    run_implemented_week2_analytics,
    tiny_avazu_dataframe,
    tiny_hillstrom_dataframe,
)
from paths import SQL_DIR

pytestmark = [pytest.mark.unit, pytest.mark.duckdb, pytest.mark.exports, pytest.mark.excel]

EXPECTED_CSV_FILES = (
    "campaign_kpis.csv",
    "ctr_trends.csv",
    "segment_performance.csv",
    "ab_test_results.csv",
    "forecast_results.csv",
    "recommendation_matrix.csv",
)

EXPECTED_EXCEL_SHEETS = (
    "Campaign_KPIs",
    "CTR_Trends",
    "Segment_Performance",
    "AB_Test_Results",
    "Forecast_Results",
    "Recommendations",
    "AB_Calculator",
)


def _build_full_tiny_bundle(tmp_path):
    raw_dir = tmp_path / "data" / "raw"
    processed = tmp_path / "data" / "processed"
    marts = tmp_path / "data" / "marts"
    exports = tmp_path / "data" / "exports"
    docs = tmp_path / "docs"
    excel_dir = tmp_path / "excel"
    raw_dir.mkdir(parents=True)
    processed.mkdir(parents=True)
    marts.mkdir(parents=True)
    exports.mkdir(parents=True)
    docs.mkdir(parents=True)
    excel_dir.mkdir(parents=True)

    avazu_raw = raw_dir / "avazu_train.csv"
    hillstrom_raw = raw_dir / "hillstrom_email.csv"
    tiny_avazu_dataframe().to_csv(avazu_raw, index=False)
    tiny_hillstrom_dataframe().to_csv(hillstrom_raw, index=False)

    avazu_clean, _ = clean_avazu_ads(pd.read_csv(avazu_raw))
    hillstrom_clean, _ = clean_hillstrom_email(pd.read_csv(hillstrom_raw))
    avazu_parquet = processed / "avazu_clean.parquet"
    hillstrom_parquet = processed / "hillstrom_clean.parquet"
    avazu_clean.to_parquet(avazu_parquet, index=False)
    hillstrom_clean.to_parquet(hillstrom_parquet, index=False)

    config = db_setup.DatabaseConfig(database_path=processed / "marketing_analytics.duckdb")
    db_setup.create_database(config=config, sql_dir=SQL_DIR)
    loader.load_data(
        config=config,
        targets=loader.get_load_targets(
            avazu_raw=avazu_raw,
            hillstrom_raw=hillstrom_raw,
            avazu_clean=avazu_parquet,
            hillstrom_clean=hillstrom_parquet,
        ),
        summary_path=processed / "duckdb_load_summary.json",
    )
    run_implemented_week2_analytics(config, processed)
    recommendations.generate_recommendations(
        config=config,
        recommendations_path=docs / "recommendations.md",
        executive_path=docs / "executive_summary.md",
        summary_path=processed / "recommendations_summary.json",
    )
    return {
        "config": config,
        "processed": processed,
        "marts": marts,
        "exports": exports,
        "excel_path": excel_dir / "marketing_executive_workbook.xlsx",
    }


def test_export_dashboard_module_exports_main():
    assert hasattr(dashboard_exports, "main")
    assert hasattr(dashboard_exports, "export_dashboard_data")


def test_export_dashboard_data_writes_csv_excel_and_summary(tmp_path):
    bundle = _build_full_tiny_bundle(tmp_path)
    summary_path = bundle["processed"] / "export_dashboard_summary.json"

    summary = dashboard_exports.export_dashboard_data(
        config=bundle["config"],
        marts_dir=bundle["marts"],
        exports_dir=bundle["exports"],
        excel_path=bundle["excel_path"],
        recommendations_summary_path=bundle["processed"] / "recommendations_summary.json",
        summary_path=summary_path,
        upload_to_s3=False,
    )

    assert summary["success"] is True
    assert summary["export_count"] == len(EXPECTED_CSV_FILES)
    assert summary_path.exists()

    for csv_name in EXPECTED_CSV_FILES:
        marts_file = bundle["marts"] / csv_name
        exports_file = bundle["exports"] / csv_name
        assert marts_file.is_file()
        assert exports_file.is_file()
        assert marts_file.read_text(encoding="utf-8") == exports_file.read_text(encoding="utf-8")

    manifest_path = bundle["exports"] / dashboard_exports.TABLEAU_MANIFEST_JSON
    assert manifest_path.is_file()
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert len(manifest["files"]) == len(EXPECTED_CSV_FILES)

    workbook = load_workbook(bundle["excel_path"], read_only=True)
    try:
        assert workbook.sheetnames == list(EXPECTED_EXCEL_SHEETS)
    finally:
        workbook.close()


def test_recommendation_matrix_csv_has_action_column(tmp_path):
    bundle = _build_full_tiny_bundle(tmp_path)
    dashboard_exports.export_dashboard_data(
        config=bundle["config"],
        marts_dir=bundle["marts"],
        exports_dir=bundle["exports"],
        excel_path=bundle["excel_path"],
        recommendations_summary_path=bundle["processed"] / "recommendations_summary.json",
        summary_path=bundle["processed"] / "export_dashboard_summary.json",
    )

    frame = pd.read_csv(bundle["marts"] / "recommendation_matrix.csv")
    assert "action" in frame.columns
    assert len(frame) >= 1
    assert set(frame["action"]).issubset({"Scale", "Pause", "Retest"})


def test_export_dashboard_data_fails_without_recommendations_summary(tmp_path):
    bundle = _build_full_tiny_bundle(tmp_path)
    rec_path = bundle["processed"] / "recommendations_summary.json"
    rec_path.unlink()

    with pytest.raises(FileNotFoundError, match="recommendations summary"):
        dashboard_exports.export_dashboard_data(
            config=bundle["config"],
            marts_dir=bundle["marts"],
            exports_dir=bundle["exports"],
            excel_path=bundle["excel_path"],
            recommendations_summary_path=rec_path,
            summary_path=bundle["processed"] / "export_dashboard_summary.json",
        )


def test_export_dashboard_data_fails_without_marts(tmp_path):
    processed = tmp_path / "processed"
    processed.mkdir()
    config = db_setup.DatabaseConfig(database_path=processed / "empty.duckdb")
    db_setup.create_database(config=config, sql_dir=SQL_DIR)

    rec_summary = processed / "recommendations_summary.json"
    rec_summary.write_text(
        json.dumps({"recommendations": [{"action": "Scale"}]}),
        encoding="utf-8",
    )

    with pytest.raises(RuntimeError, match="not fully populated"):
        dashboard_exports.export_dashboard_data(
            config=config,
            marts_dir=tmp_path / "marts",
            exports_dir=tmp_path / "exports",
            excel_path=tmp_path / "workbook.xlsx",
            recommendations_summary_path=rec_summary,
            summary_path=processed / "export_dashboard_summary.json",
        )


def test_ab_calculator_sheet_contains_lift_formulas(tmp_path):
    bundle = _build_full_tiny_bundle(tmp_path)
    dashboard_exports.export_dashboard_data(
        config=bundle["config"],
        marts_dir=bundle["marts"],
        exports_dir=bundle["exports"],
        excel_path=bundle["excel_path"],
        recommendations_summary_path=bundle["processed"] / "recommendations_summary.json",
        summary_path=bundle["processed"] / "export_dashboard_summary.json",
    )

    workbook = load_workbook(bundle["excel_path"], data_only=False)
    try:
        sheet = workbook["AB_Calculator"]
        assert sheet["C15"].value == "=C14-B14"
        assert sheet["D16"].value == '=IF(B14=0,"",(D14-B14)/B14)'
    finally:
        workbook.close()
