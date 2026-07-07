#!/usr/bin/env python3
"""Export DuckDB marts to CSV, Excel workbook, and optional S3."""

from __future__ import annotations

import json
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from create_duckdb_database import DatabaseConfig, _display_path, load_config
from paths import (
    EXCEL_WORKBOOK,
    EXPORT_DASHBOARD_SUMMARY,
    EXPORTS_DIR,
    MARTS_DIR,
    PROJECT_ROOT,
    RECOMMENDATIONS_SUMMARY,
)
from upload_to_s3 import (
    UploadConfig,
    UploadTarget,
    create_s3_client,
    load_config as load_s3_config,
    upload_targets,
    validate_local_files,
    verify_bucket_access,
)

REQUIRED_MART_TABLES = (
    "mart_campaign_kpis",
    "mart_ctr_trends",
    "mart_device_app_performance",
    "mart_ab_test_results",
    "mart_forecast_inputs",
    "mart_forecast_results",
)

MART_CSV_EXPORTS: tuple[tuple[str, str, str], ...] = (
    ("mart_campaign_kpis", "campaign_kpis.csv", "Campaign KPIs"),
    ("mart_ctr_trends", "ctr_trends.csv", "CTR trends"),
    (
        "mart_device_app_performance",
        "segment_performance.csv",
        "Segment performance",
    ),
    ("mart_ab_test_results", "ab_test_results.csv", "A/B test results"),
    ("mart_forecast_results", "forecast_results.csv", "Forecast results"),
)

RECOMMENDATION_MATRIX_CSV = "recommendation_matrix.csv"
TABLEAU_MANIFEST_JSON = "tableau_data_manifest.json"


@dataclass(frozen=True)
class ExportRecord:
    source_table: str
    csv_name: str
    description: str
    row_count: int
    marts_path: Path
    exports_path: Path


def ensure_database_ready(config: DatabaseConfig) -> None:
    if not config.database_path.exists():
        raise RuntimeError(
            f"DuckDB database not found at {config.database_path}. "
            "Run the analytics pipeline before exporting dashboard data."
        )


def _mart_has_rows(connection: duckdb.DuckDBPyConnection, table_name: str) -> bool:
    count = connection.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    return int(count) > 0


def ensure_marts_populated(connection: duckdb.DuckDBPyConnection) -> None:
    missing = [
        table_name
        for table_name in REQUIRED_MART_TABLES
        if not _mart_has_rows(connection, table_name)
    ]
    if missing:
        raise RuntimeError(
            "Analytics marts are not fully populated. Missing or empty tables:\n  - "
            + "\n  - ".join(missing)
        )


def load_recommendations_summary(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(
            f"Missing recommendations summary at {path}. "
            "Run `python scripts/generate_recommendations.py` first."
        )
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not payload.get("recommendations"):
        raise RuntimeError("Recommendations summary has no recommendation rows.")
    return payload


def export_mart_table(
    connection: duckdb.DuckDBPyConnection,
    table_name: str,
    csv_name: str,
    description: str,
    marts_dir: Path,
    exports_dir: Path,
) -> ExportRecord:
    frame = connection.execute(
        f"""
        SELECT * EXCLUDE (created_at)
        FROM {table_name}
        """
    ).fetchdf()

    marts_dir.mkdir(parents=True, exist_ok=True)
    exports_dir.mkdir(parents=True, exist_ok=True)

    marts_path = marts_dir / csv_name
    exports_path = exports_dir / csv_name
    frame.to_csv(marts_path, index=False)
    shutil.copy2(marts_path, exports_path)

    return ExportRecord(
        source_table=table_name,
        csv_name=csv_name,
        description=description,
        row_count=len(frame),
        marts_path=marts_path,
        exports_path=exports_path,
    )


def export_recommendation_matrix(
    recommendations_summary: dict[str, Any],
    marts_dir: Path,
    exports_dir: Path,
) -> ExportRecord:
    frame = pd.DataFrame(recommendations_summary["recommendations"])
    column_order = ["channel", "segment", "metric", "action", "evidence", "caveat"]
    frame = frame.reindex(columns=column_order)

    marts_dir.mkdir(parents=True, exist_ok=True)
    exports_dir.mkdir(parents=True, exist_ok=True)

    marts_path = marts_dir / RECOMMENDATION_MATRIX_CSV
    exports_path = exports_dir / RECOMMENDATION_MATRIX_CSV
    frame.to_csv(marts_path, index=False)
    shutil.copy2(marts_path, exports_path)

    return ExportRecord(
        source_table="recommendations_summary",
        csv_name=RECOMMENDATION_MATRIX_CSV,
        description="Scale / pause / retest recommendation matrix",
        row_count=len(frame),
        marts_path=marts_path,
        exports_path=exports_path,
    )


def _write_dataframe_sheet(workbook: Workbook, sheet_name: str, frame: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    for row in dataframe_to_rows(frame, index=False, header=True):
        worksheet.append(row)
    _format_header_row(worksheet)
    _autosize_columns(worksheet)
    worksheet.freeze_panes = "A2"


def _format_header_row(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(worksheet, max_width: int = 42) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max_length + 2, max_width)


def _build_executive_summary_sheet(
    workbook: Workbook,
    campaign_frame: pd.DataFrame,
    recommendations_frame: pd.DataFrame,
) -> None:
    worksheet = workbook.create_sheet(title="Executive_Summary", index=0)
    bold = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14, color="1F4E79")

    row = campaign_frame.iloc[0]
    action_counts = recommendations_frame["action"].value_counts()

    worksheet["A1"] = "Marketing Executive Workbook — Summary"
    worksheet["A1"].font = title_font
    worksheet["A3"] = "Metric"
    worksheet["B3"] = "Value"
    worksheet["A3"].font = bold
    worksheet["B3"].font = bold

    summary_rows = [
        ("Impressions", f"{int(row['impressions']):,}"),
        ("Clicks", f"{int(row['clicks']):,}"),
        ("Portfolio CTR", f"{float(row['ctr']) * 100:.2f}%"),
        ("Recommendations", str(len(recommendations_frame))),
        ("Scale actions", str(int(action_counts.get("Scale", 0)))),
        ("Pause actions", str(int(action_counts.get("Pause", 0)))),
        ("Retest actions", str(int(action_counts.get("Retest", 0)))),
    ]
    for offset, (label, value) in enumerate(summary_rows, start=4):
        worksheet.cell(row=offset, column=1, value=label)
        worksheet.cell(row=offset, column=2, value=value)

    worksheet["A12"] = "Notes"
    worksheet["A12"].font = bold
    worksheet["A13"] = (
        "Stakeholder workbook built from DuckDB mart exports. "
        "Forecast MAPE is directional only on single-day Avazu data."
    )
    _autosize_columns(worksheet)


def _build_pivot_recommendations_sheet(workbook: Workbook, recommendations_frame: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title="Pivot_Recommendations")
    pivot = (
        recommendations_frame.groupby("action", as_index=False)
        .size()
        .rename(columns={"size": "recommendation_count"})
        .sort_values("action")
    )

    for row in dataframe_to_rows(pivot, index=False, header=True):
        worksheet.append(row)
    _format_header_row(worksheet)
    _autosize_columns(worksheet)

    chart = BarChart()
    chart.title = "Recommendations by Action"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Action"
    data = Reference(worksheet, min_col=2, min_row=1, max_row=len(pivot) + 1)
    categories = Reference(worksheet, min_col=1, min_row=2, max_row=len(pivot) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 14
    worksheet.add_chart(chart, "D2")


def _build_ab_calculator_sheet(workbook: Workbook, ab_frame: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title="AB_Calculator")
    bold = Font(bold=True)

    worksheet["A1"] = "Hillstrom A/B scenario calculator"
    worksheet["A1"].font = bold
    worksheet["A3"] = "Field"
    worksheet["B3"] = "Control"
    worksheet["C3"] = "Mens E-Mail"
    worksheet["D3"] = "Womens E-Mail"
    for cell in worksheet[3]:
        cell.font = bold

    control = ab_frame[ab_frame["treatment_group"] == "control"].iloc[0]
    mens = ab_frame[ab_frame["treatment_group"] == "mens_email"].iloc[0]
    womens = ab_frame[ab_frame["treatment_group"] == "womens_email"].iloc[0]

    rows = [
        ("Recipients", int(control["recipients"]), int(mens["recipients"]), int(womens["recipients"])),
        ("Conversions", int(control["conversions"]), int(mens["conversions"]), int(womens["conversions"])),
        (
            "Conversion rate",
            float(control["conversion_rate"]),
            float(mens["conversion_rate"]),
            float(womens["conversion_rate"]),
        ),
        (
            "Absolute lift vs control",
            0.0,
            float(mens["absolute_lift"]),
            float(womens["absolute_lift"]),
        ),
        (
            "Relative lift %",
            0.0,
            float(mens["relative_lift_pct"]),
            float(womens["relative_lift_pct"]),
        ),
        (
            "Incremental revenue",
            0.0,
            float(mens["incremental_revenue"]),
            float(womens["incremental_revenue"]),
        ),
        (
            "p-value vs control",
            "",
            float(mens["p_value"]) if mens["p_value"] is not None else "",
            float(womens["p_value"]) if womens["p_value"] is not None else "",
        ),
    ]

    start_row = 4
    for offset, (label, control_value, mens_value, womens_value) in enumerate(rows):
        row_idx = start_row + offset
        worksheet.cell(row=row_idx, column=1, value=label)
        worksheet.cell(row=row_idx, column=2, value=control_value)
        worksheet.cell(row=row_idx, column=3, value=mens_value)
        worksheet.cell(row=row_idx, column=4, value=womens_value)

    worksheet["A12"] = "Scenario inputs (edit conversion rates to model lift)"
    worksheet["A12"].font = bold
    worksheet["A14"] = "Scenario conversion rate"
    worksheet["B14"] = float(control["conversion_rate"])
    worksheet["C14"] = float(mens["conversion_rate"])
    worksheet["D14"] = float(womens["conversion_rate"])
    worksheet["A15"] = "Scenario absolute lift vs control"
    worksheet["B15"] = 0
    worksheet["C15"] = "=C14-B14"
    worksheet["D15"] = "=D14-B14"
    worksheet["A16"] = "Scenario relative lift %"
    worksheet["B16"] = 0
    worksheet["C16"] = "=IF(B14=0,\"\",(C14-B14)/B14)"
    worksheet["D16"] = "=IF(B14=0,\"\",(D14-B14)/B14)"

    _autosize_columns(worksheet)
    for row_idx in range(4, 11):
        for col_idx in range(2, 5):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, float) and row_idx in {6, 7, 8}:
                cell.number_format = "0.00%"
            elif isinstance(cell.value, float) and row_idx in {5, 10}:
                cell.number_format = "#,##0"
    for col_idx in range(2, 5):
        worksheet.cell(row=14, column=col_idx).number_format = "0.00%"
        worksheet.cell(row=15, column=col_idx).number_format = "0.00%"
        worksheet.cell(row=16, column=col_idx).number_format = "0.00%"


def build_excel_workbook(
    frames: dict[str, pd.DataFrame],
    recommendations_summary: dict[str, Any],
    excel_path: Path,
) -> Path:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    recommendations_frame = pd.DataFrame(recommendations_summary["recommendations"])

    _build_executive_summary_sheet(workbook, frames["campaign_kpis"], recommendations_frame)
    _write_dataframe_sheet(workbook, "Campaign_KPIs", frames["campaign_kpis"])
    _write_dataframe_sheet(workbook, "CTR_Trends", frames["ctr_trends"])
    _write_dataframe_sheet(workbook, "Segment_Performance", frames["segment_performance"])
    _write_dataframe_sheet(workbook, "AB_Test_Results", frames["ab_test_results"])
    _write_dataframe_sheet(workbook, "Forecast_Results", frames["forecast_results"])
    _write_dataframe_sheet(workbook, "Recommendations", recommendations_frame)
    _build_pivot_recommendations_sheet(workbook, recommendations_frame)
    _build_ab_calculator_sheet(workbook, frames["ab_test_results"])

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)
    return excel_path


def build_tableau_manifest(exports: list[ExportRecord], exports_dir: Path) -> Path:
    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "purpose": "Tableau dashboard data connections",
        "files": [
            {
                "file": record.csv_name,
                "source": record.source_table,
                "description": record.description,
                "row_count": record.row_count,
                "relative_path": _display_path(record.exports_path),
            }
            for record in exports
        ],
    }
    manifest_path = exports_dir / TABLEAU_MANIFEST_JSON
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return manifest_path


def get_s3_export_targets(
    config: UploadConfig,
    exports: list[ExportRecord],
    excel_path: Path,
    manifest_path: Path,
) -> list[UploadTarget]:
    targets = [
        UploadTarget(
            local_path=record.marts_path,
            s3_key=f"{config.marts_prefix}/{record.csv_name}",
        )
        for record in exports
    ]
    targets.append(
        UploadTarget(
            local_path=excel_path,
            s3_key=f"{config.export_prefix}/marketing_executive_workbook.xlsx",
        )
    )
    targets.append(
        UploadTarget(
            local_path=manifest_path,
            s3_key=f"{config.export_prefix}/{TABLEAU_MANIFEST_JSON}",
        )
    )
    for record in exports:
        targets.append(
            UploadTarget(
                local_path=record.exports_path,
                s3_key=f"{config.export_prefix}/{record.csv_name}",
            )
        )
    return targets


def export_dashboard_data(
    config: DatabaseConfig | None = None,
    marts_dir: Path | None = None,
    exports_dir: Path | None = None,
    excel_path: Path | None = None,
    recommendations_summary_path: Path | None = None,
    summary_path: Path | None = None,
    upload_to_s3: bool = False,
) -> dict[str, Any]:
    config = config or load_config()
    marts_dir = marts_dir or MARTS_DIR
    exports_dir = exports_dir or EXPORTS_DIR
    excel_path = excel_path or EXCEL_WORKBOOK
    recommendations_summary_path = recommendations_summary_path or RECOMMENDATIONS_SUMMARY
    summary_path = summary_path or EXPORT_DASHBOARD_SUMMARY

    ensure_database_ready(config)
    recommendations_summary = load_recommendations_summary(recommendations_summary_path)

    connection = duckdb.connect(str(config.database_path), read_only=True)
    try:
        ensure_marts_populated(connection)
        exports: list[ExportRecord] = []
        frames: dict[str, pd.DataFrame] = {}

        for table_name, csv_name, description in MART_CSV_EXPORTS:
            record = export_mart_table(
                connection,
                table_name,
                csv_name,
                description,
                marts_dir=marts_dir,
                exports_dir=exports_dir,
            )
            exports.append(record)
            key = csv_name.removesuffix(".csv")
            frames[key] = pd.read_csv(record.marts_path)

        recommendation_export = export_recommendation_matrix(
            recommendations_summary,
            marts_dir=marts_dir,
            exports_dir=exports_dir,
        )
        exports.append(recommendation_export)
    finally:
        connection.close()

    excel_output = build_excel_workbook(frames, recommendations_summary, excel_path)
    manifest_path = build_tableau_manifest(exports, exports_dir)

    s3_summary: dict[str, Any] | None = None
    if upload_to_s3:
        s3_config = load_s3_config()
        targets = get_s3_export_targets(s3_config, exports, excel_output, manifest_path)
        validate_local_files(targets)
        client = create_s3_client(s3_config)
        verify_bucket_access(client, s3_config.bucket)
        s3_summary = upload_targets(client, s3_config, targets)

    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "database_path": _display_path(config.database_path),
        "marts_dir": _display_path(marts_dir),
        "exports_dir": _display_path(exports_dir),
        "excel_workbook": _display_path(excel_output),
        "tableau_manifest": _display_path(manifest_path),
        "csv_exports": [
            {
                "source_table": record.source_table,
                "csv_name": record.csv_name,
                "description": record.description,
                "row_count": record.row_count,
                "marts_path": _display_path(record.marts_path),
                "exports_path": _display_path(record.exports_path),
            }
            for record in exports
        ],
        "export_count": len(exports),
        "recommendation_count": len(recommendations_summary["recommendations"]),
        "s3_upload": s3_summary,
        "success": True,
    }

    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    return summary


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Export marts for Tableau and Excel.")
    parser.add_argument(
        "--upload-s3",
        action="store_true",
        help="Upload mart CSVs, exports, and Excel workbook to S3 after local export.",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("Tableau / Excel dashboard exports")
    print("=" * 60)

    try:
        summary = export_dashboard_data(upload_to_s3=args.upload_s3)
        print(f"Wrote {len(summary['csv_exports'])} CSV exports to {summary['marts_dir']}")
        print(f"Copied exports to {summary['exports_dir']}")
        print(f"Wrote Excel workbook: {summary['excel_workbook']}")
        print(f"Wrote Tableau manifest: {summary['tableau_manifest']}")
        print(f"Wrote {EXPORT_DASHBOARD_SUMMARY}")
        if args.upload_s3 and summary.get("s3_upload"):
            s3_summary = summary["s3_upload"]
            print(
                f"Uploaded {s3_summary['uploaded_count']} of "
                f"{len(s3_summary['uploads'])} files to s3://{s3_summary['bucket']}"
            )
            if not s3_summary["success"]:
                return 1
        return 0
    except (RuntimeError, FileNotFoundError, ValueError) as exc:
        print(f"Export failed: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
